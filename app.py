import os, io, json, re, warnings
import numpy as np
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, render_template, send_file
import plotly.graph_objects as go

warnings.filterwarnings("ignore")

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── colours matching reference images ──────────────────────────────────────────
C_PREV   = "#4a9fd4"   # light blue  — previous year
C_CURR   = "#1a3c5e"   # dark navy   — current year
C_UP     = "#27ae60"   # green ▲
C_DOWN   = "#e74c3c"   # red ▼
C_GRID   = "#e8ecf0"
C_BG     = "#ffffff"
FONT     = "Inter, -apple-system, BlinkMacSystemFont, Segoe UI, sans-serif"

# ── category keyword map ────────────────────────────────────────────────────────
CATEGORY_MAP = {
    "Beer": {
        "kw": ["beer","pils","lager","ale","ipa","stout","porter","brew","бира",
                "chimay","westmalle","leffe","kwaremont","bavik","lindemans","corona",
                "heineken","stella","budweiser","petrus","bernardus","wittekerke","draft"],
        "sub": {
            "Craft Beer": ["chimay","westmalle","leffe","kwaremont","bernardus","wittekerke",
                           "lindemans","petrus","craft","abbey","trappist"],
            "Draft Beer":  ["draft"],
            "Guest Beer":  ["guest","import","corona","heineken","stella","budweiser"],
        },
    },
    "Wine": {
        "kw": ["wine","дарс","merlot","chardonnay","sauvignon","champagne","prosecco",
                "moet","ros","shiraz","cabernet","pinot","riesling","шампань","bardolino",
                "varaschin","cormons","poesie","terre forti","toso","moscato","trebbino"],
        "sub": {
            "Red Wine":    ["merlot","shiraz","cabernet","pinot noir","bardolino","улаан"],
            "White Wine":  ["chardonnay","sauvignon","riesling","pinot grigio","цагаан",
                            "cormons","moscato","trebbino","varaschin"],
            "Champagne":   ["champagne","prosecco","moet","sparkling","шампань","cava"],
            "Rosé":        ["ros","ягаан"],
        },
    },
    "Spirits": {
        "kw": ["vodka","gin","whiskey","whisky","cognac","brandy","tequila","rum",
                "scotch","bourbon","liqueur","absolut","grey goose","greygoose","beluga",
                "finlandia","soyombo","bombay","hendricks","gordons","gordon","tanqueray",
                "hennessy","jameson","patron","jose cuervo","sierra","glenmorange",
                "single malt","set","архи","спирт"],
        "sub": {
            "Vodka":    ["vodka","beluga","finlandia","grey goose","greygoose","absolut","soyombo"],
            "Gin":      ["gin","bombay","hendricks","gordons","gordon","tanqueray"],
            "Whiskey":  ["whiskey","whisky","jameson","scotch","bourbon","glenmorange","single malt"],
            "Cognac":   ["cognac","hennessy","brandy"],
            "Tequila":  ["tequila","patron","jose cuervo","sierra"],
            "Rum":      ["rum","bacardi"],
        },
    },
    "Cocktails": {
        "kw": ["cocktail","mojito","margarita","martini","negroni","sour","highball",
                "punch","belgian long island","pink dream","red diamond","tom collins",
                "коктейль","lemonade","лимонад","blackcurrant","mango lemon","strawberry lemon",
                "shake","cinamon"],
        "sub": {
            "Cocktails":  ["negroni","martini","manhattan","old fashioned","daiquiri",
                           "belgian long island","pink dream","red diamond","tom collins",
                           "whiskey sour","hennessy highball","special cocktail"],
            "Lemonade":   ["lemonade","лимонад","blackcurrant","mango lemon","strawberry lemon",
                           "punch","cinamon"],
            "Milkshake":  ["shake","milkshake"],
        },
    },
    "Soft Drinks": {
        "kw": ["soda","cola","pepsi","sprite","fanta","juice","water","ус","tonic",
                "ginger beer","redbull","red bull","schweppes","khujirt","millenia",
                "club soda","soft drink","langers","ice tea","iced tea","lemon tea",
                "шүүс","premium"],
        "sub": {
            "Water":       ["water","ус","khujirt","millenia"],
            "Soft Drinks": ["soda","cola","pepsi","sprite","fanta","schweppes","tonic",
                            "ginger beer","club soda","premium","soft drink"],
            "Juice":       ["juice","шүүс","langers","apple","mango","orange","ice tea","lemon tea"],
            "Energy":      ["redbull","red bull","energy"],
        },
    },
    "Hot Drinks": {
        "kw": ["tea","coffee","цай","кофе","espresso","latte","cappuccino","americano",
                "hot chocolate","althaus","халуун"],
        "sub": {
            "Tea":       ["tea","цай","althaus","green","black","berry"],
            "Coffee":    ["coffee","кофе","espresso","latte","cappuccino","americano"],
            "Other Hot": ["hot chocolate","халуун ундаа"],
        },
    },
    "Food": {
        "kw": ["pizza","steak","chicken","pork","beef","salmon","fish","pasta","fries",
                "salad","soup","wings","burger","platter","appetizer","starter","dessert",
                "хоол","пицца","goulash","carbonara","spareribs","margherita","meat lovers",
                "spicy salami","penne","ox bone","pork belly","chicken leg","fillet",
                "cobb","avocado","cheese","sausage","french fries","bacon","snack","grill"],
        "sub": {
            "Pizza":            ["pizza","margherita","meat lovers","spicy salami","пицца"],
            "Main Course":      ["steak","salmon","pork chop","pork belly","bbq","spareribs",
                                 "chicken leg","goulash","carbonara","penne","ox bone","fillet",
                                 "үндсэн"],
            "Starters & Sides": ["fries","wings","snack","platter","salad","cheese","sausage",
                                  "appetizer","starter","cobb","avocado","bacon","grill","share"],
            "Pasta":            ["pasta","carbonara","penne"],
            "Pub Food":         ["burger","sandwich","pub food"],
        },
    },
    "Karaoke & Shisha": {
        "kw": ["karaoke","shisha","hookah","карааке","шиша"],
        "sub": {
            "Karaoke": ["karaoke","карааке"],
            "Shisha":  ["shisha","hookah","шиша"],
        },
    },
}

# ── helpers ─────────────────────────────────────────────────────────────────────

def coerce_numerics(df):
    for col in df.columns:
        if df[col].dtype == object:
            cv = pd.to_numeric(df[col], errors="coerce")
            if cv.notna().mean() > 0.6:
                df[col] = cv
    return df


def detect_cols(df):
    """Return best-guess column names for date, item, qty, revenue, staff."""
    cols = {c.lower(): c for c in df.columns}
    def pick(*hints):
        for h in hints:
            for k, v in cols.items():
                if h in k:
                    return v
        return None

    date_col = pick("огноо","date","дата","datetime","time","өдөр")
    item_col = pick("бараа нэр","нэр","бараа","item","product","goods","name","барааны")
    qty_col  = pick("тоо","qty","quantity","count","ширхэг")
    rev_col  = pick("нийт дүн","дүн","total","revenue","орлого","sum")
    emp_col  = pick("сонгогдсон","ажилтан","staff","employee","waiter","cashier","хүн")
    cat_col  = pick("ангилал","category","type","бүлэг","group")

    # fallback: pick largest-sum numeric for revenue
    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if rev_col is None and num_cols:
        rev_col = max(num_cols, key=lambda c: df[c].sum(skipna=True))
    if qty_col is None and num_cols:
        others = [c for c in num_cols if c != rev_col]
        qty_col = others[0] if others else rev_col

    return date_col, item_col, qty_col, rev_col, emp_col, cat_col


def assign_category(name: str):
    """Return (category, subcategory) for a product name string."""
    n = str(name).lower()
    for cat, meta in CATEGORY_MAP.items():
        if any(kw in n for kw in meta["kw"]):
            for sub, skws in meta["sub"].items():
                if any(kw in n for kw in skws):
                    return cat, sub
            return cat, cat   # no sub match → sub = cat itself
    return "Other", "Other"


def fmt(v):
    if pd.isna(v) or v == 0:
        return "0"
    av = abs(v)
    if av >= 1_000_000:
        return f"{v/1_000_000:.1f}M"
    if av >= 1_000:
        return f"{v/1_000:.0f}K"
    return f"{v:.0f}"


def pct_change(prev, curr):
    if prev and prev != 0:
        return (curr - prev) / abs(prev) * 100
    return 100.0 if curr else 0.0


def clear_uploads():
    i = 0
    while True:
        p = os.path.join(UPLOAD_FOLDER, f"upload_{i}.xlsx")
        if not os.path.exists(p):
            break
        os.remove(p)
        np = p.replace(".xlsx", ".name")
        if os.path.exists(np):
            os.remove(np)
        i += 1


def to_json(fig):
    return json.loads(fig.to_json())


def base_layout(**kw):
    d = dict(
        paper_bgcolor=C_BG, plot_bgcolor=C_BG,
        font=dict(family=FONT, color="#1e293b", size=12),
        title_font=dict(family=FONT, size=14, color="#0f172a"),
        legend=dict(
            bgcolor="rgba(255,255,255,0.9)",
            bordercolor=C_GRID, borderwidth=1,
            font=dict(size=11),
        ),
        hoverlabel=dict(bgcolor="white", bordercolor=C_GRID,
                        font=dict(size=11, family=FONT)),
        margin=dict(t=90, b=60, l=60, r=40),
    )
    d.update(kw)
    return d


# ── chart builders ──────────────────────────────────────────────────────────────

def yoy_vertical(prev_vals, curr_vals, labels, title, subtitle,
                 metric_label, prev_lbl, curr_lbl, source_label="", top=20):
    """Grouped vertical bar with % change annotations — mirrors reference overview charts."""
    # sort by current value desc, cap at top
    paired = sorted(zip(curr_vals, prev_vals, labels), reverse=True)[:top]
    curr_vals, prev_vals, labels = zip(*paired) if paired else ([], [], [])
    curr_vals, prev_vals, labels = list(curr_vals), list(prev_vals), list(labels)

    annotations = []
    for i, (p, c) in enumerate(zip(prev_vals, curr_vals)):
        pct = pct_change(p, c)
        col = C_UP if pct >= 0 else C_DOWN
        arrow = "▲" if pct >= 0 else "▼"
        y_pos = max(p, c)
        annotations.append(dict(
            x=labels[i], y=y_pos,
            xref="x", yref="y",
            text=f"<b>{arrow}{abs(pct):.0f}%</b>",
            showarrow=False, yshift=10,
            font=dict(size=11, family=FONT, color=col),
            align="center",
        ))

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name=prev_lbl, x=labels, y=prev_vals,
        marker=dict(color=C_PREV, line=dict(width=0)),
        text=[fmt(v) for v in prev_vals], textposition="outside",
        textfont=dict(size=10, color="#475569"),
        hovertemplate="<b>%{x}</b><br>" + prev_lbl + ": %{y:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name=curr_lbl, x=labels, y=curr_vals,
        marker=dict(color=C_CURR, line=dict(width=0)),
        text=[fmt(v) for v in curr_vals], textposition="outside",
        textfont=dict(size=10, color="#0f172a"),
        hovertemplate="<b>%{x}</b><br>" + curr_lbl + ": %{y:,.0f}<extra></extra>",
    ))

    full_title = f"<b>{title}</b><br><sup>{subtitle}</sup>"
    height = max(480, len(labels) * 18 + 180)

    layout = base_layout(
        barmode="group",
        title=dict(text=full_title, x=0.5, xanchor="center", y=0.97),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=11)),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title=metric_label,
                   tickfont=dict(size=10)),
        annotations=annotations,
        height=height,
        margin=dict(t=100, b=70, l=70, r=40),
        legend=dict(orientation="v", x=1.01, y=0.99, xanchor="left"),
    )
    if source_label:
        layout["annotations"] = layout.get("annotations", []) + [dict(
            text=f"<b>● {source_label}</b>",
            x=0, y=1.08, xref="paper", yref="paper",
            showarrow=False, font=dict(size=11, color=C_CURR),
            align="left",
        )]
    fig.update_layout(**layout)
    return fig


def yoy_horizontal(prev_vals, curr_vals, labels, title, subtitle,
                    metric_label, prev_lbl, curr_lbl, source_label=""):
    """Grouped horizontal bar with % on right — mirrors reference item-detail charts."""
    # sort by current value asc (so biggest is at top)
    paired = sorted(zip(curr_vals, prev_vals, labels))
    curr_vals, prev_vals, labels = zip(*paired) if paired else ([], [], [])
    curr_vals, prev_vals, labels = list(curr_vals), list(prev_vals), list(labels)

    annotations = []
    max_x = max(list(curr_vals) + list(prev_vals), default=1)
    for i, (p, c) in enumerate(zip(prev_vals, curr_vals)):
        pct = pct_change(p, c)
        col = C_UP if pct >= 0 else C_DOWN
        arrow = "▲" if pct >= 0 else "▼"
        annotations.append(dict(
            x=max_x * 1.02, y=labels[i],
            xref="x", yref="y",
            text=f"<b>{arrow}{abs(pct):.0f}%</b>",
            showarrow=False, xanchor="left",
            font=dict(size=10, family=FONT, color=col),
        ))

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name=prev_lbl, y=labels, x=prev_vals, orientation="h",
        marker=dict(color=C_PREV, line=dict(width=0)),
        text=[fmt(v) for v in prev_vals], textposition="outside",
        textfont=dict(size=10, color="#475569"),
        hovertemplate="<b>%{y}</b><br>" + prev_lbl + ": %{x:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name=curr_lbl, y=labels, x=curr_vals, orientation="h",
        marker=dict(color=C_CURR, line=dict(width=0)),
        text=[fmt(v) for v in curr_vals], textposition="outside",
        textfont=dict(size=10, color="#0f172a"),
        hovertemplate="<b>%{y}</b><br>" + curr_lbl + ": %{x:,.0f}<extra></extra>",
    ))

    full_title = f"<b>{title}</b><br><sup>{subtitle}</sup>"
    height = max(420, len(labels) * 42 + 120)

    layout = base_layout(
        barmode="group",
        title=dict(text=full_title, x=0.5, xanchor="center", y=0.97),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title=metric_label,
                   tickfont=dict(size=10), range=[0, max_x * 1.25]),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=11)),
        annotations=annotations,
        height=height,
        margin=dict(t=100, b=60, l=160, r=80),
        legend=dict(orientation="v", x=1.0, y=0.01, xanchor="left"),
    )
    if source_label:
        layout["annotations"] = layout.get("annotations", []) + [dict(
            text=f"<b>● {source_label}</b>",
            x=0, y=1.10, xref="paper", yref="paper",
            showarrow=False, font=dict(size=11, color=C_CURR), align="left",
        )]
    fig.update_layout(**layout)
    return fig


def staff_revenue_chart(curr_df, emp_col, rev_col, curr_lbl, subtitle, source_label=""):
    """Staff revenue bar with total footer — mirrors sales_karaoke.png style."""
    g = curr_df.groupby(emp_col)[rev_col].sum().sort_values(ascending=False)
    total = g.sum()

    fig = go.Figure(go.Bar(
        x=g.index.astype(str).tolist(),
        y=g.values.tolist(),
        marker=dict(color=C_CURR, line=dict(width=0)),
        text=[fmt(v) for v in g.values],
        textposition="outside",
        textfont=dict(size=12, color="#0f172a"),
        hovertemplate="<b>%{x}</b><br>%{y:,.0f}<extra></extra>",
    ))

    title_text = (f"<b>Staff Revenue — {source_label + ' | ' if source_label else ''}"
                  f"{curr_lbl}</b><br><sup>{subtitle}</sup>")

    fig.update_layout(**base_layout(
        title=dict(text=title_text, x=0.5, xanchor="center", y=0.95),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=12)),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID,
                   tickfont=dict(size=10), title="Revenue"),
        height=520,
        margin=dict(t=100, b=120, l=60, r=40),
        annotations=[
            dict(
                text=f"<b>{fmt(total)} ₮</b>",
                x=0.5, y=-0.25, xref="paper", yref="paper",
                showarrow=False, align="center",
                bgcolor=C_CURR, bordercolor=C_CURR,
                borderpad=12, borderwidth=0,
                font=dict(family=FONT, size=20, color="white"),
            ),
        ],
    ))
    if source_label:
        fig.add_annotation(
            text=f"<b>● {source_label}</b>",
            x=0, y=1.08, xref="paper", yref="paper",
            showarrow=False, font=dict(size=11, color=C_CURR), align="left",
        )
    return fig


def daily_trend_chart(prev_df, curr_df, date_col, rev_col, prev_lbl, curr_lbl, subtitle):
    """Line chart overlaying two years' daily revenue."""
    def daily(df):
        d = df.copy()
        d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
        d = d.dropna(subset=[date_col])
        d["day"] = d[date_col].dt.day
        return d.groupby("day")[rev_col].sum()

    p = daily(prev_df) if prev_df is not None else pd.Series(dtype=float)
    c = daily(curr_df)

    fig = go.Figure()
    if not p.empty:
        fig.add_trace(go.Scatter(
            x=p.index.tolist(), y=p.values.tolist(),
            name=prev_lbl, mode="lines+markers",
            line=dict(color=C_PREV, width=2.5, shape="spline", smoothing=0.7),
            marker=dict(size=5),
            hovertemplate="Day %{x}<br>" + prev_lbl + ": %{y:,.0f}<extra></extra>",
        ))
    fig.add_trace(go.Scatter(
        x=c.index.tolist(), y=c.values.tolist(),
        name=curr_lbl, mode="lines+markers",
        line=dict(color=C_CURR, width=2.5, shape="spline", smoothing=0.7),
        marker=dict(size=5),
        hovertemplate="Day %{x}<br>" + curr_lbl + ": %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(**base_layout(
        title=dict(
            text=f"<b>Daily Revenue Trend</b><br><sup>{subtitle}</sup>",
            x=0.5, xanchor="center",
        ),
        xaxis=dict(gridcolor=C_GRID, title="Day of Month"),
        yaxis=dict(gridcolor=C_GRID, title="Revenue"),
        hovermode="x unified",
        height=420,
        legend=dict(orientation="h", y=1.08, x=1, xanchor="right"),
    ))
    return fig


def dow_chart(df, date_col, rev_col, lbl, subtitle):
    """Revenue by day of week, best day highlighted."""
    d = df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    d = d.dropna(subset=[date_col])
    order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow = d.groupby(d[date_col].dt.day_name())[rev_col].sum().reindex(order).fillna(0)
    best = dow.idxmax() if dow.sum() > 0 else None
    colors = [C_CURR if d == best else C_PREV for d in order]
    annotations = []
    if best:
        annotations.append(dict(
            text=f"<b>Best: {best}</b>",
            x=0.5, y=1.05, xref="paper", yref="paper",
            showarrow=False, font=dict(size=12, color=C_CURR), align="center",
        ))
    fig = go.Figure(go.Bar(
        x=order, y=dow.values.tolist(),
        marker=dict(color=colors, line=dict(width=0)),
        text=[fmt(v) for v in dow.values], textposition="outside",
        textfont=dict(size=11, color="#0f172a"),
        hovertemplate="<b>%{x}</b><br>%{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(**base_layout(
        title=dict(text=f"<b>Revenue by Day of Week</b><br><sup>{subtitle}</sup>",
                   x=0.5, xanchor="center"),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="Revenue"),
        annotations=annotations,
        height=420,
    ))
    return fig


def donut_chart(labels, values, title):
    """Category revenue mix donut."""
    palette = ["#1a3c5e","#4a9fd4","#27ae60","#e67e22","#9b59b6",
               "#e74c3c","#1abc9c","#f39c12","#2980b9","#d35400"]
    fig = go.Figure(go.Pie(
        labels=labels, values=values,
        hole=0.45,
        textinfo="label+percent",
        textfont=dict(size=11, family=FONT),
        marker=dict(colors=palette[:len(labels)], line=dict(color="white", width=2)),
        hovertemplate="<b>%{label}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>",
        sort=True,
    ))
    fig.update_layout(**base_layout(
        title=dict(text=f"<b>{title}</b>", x=0.5, xanchor="center"),
        showlegend=True,
        legend=dict(orientation="v", x=1.01, y=0.5, xanchor="left"),
        height=480,
        margin=dict(t=80, b=40, l=40, r=120),
    ))
    return fig


def movers_chart(prev_df, curr_df, item_col, rev_col, prev_lbl, curr_lbl, subtitle):
    """Top 5 gainers + top 5 losers by % change, diverging horizontal bar."""
    curr_g = curr_df.groupby(item_col)[rev_col].sum()
    prev_g = prev_df.groupby(item_col)[rev_col].sum()
    combined = pd.DataFrame({"curr": curr_g, "prev": prev_g}).fillna(0)
    combined = combined[combined["curr"] + combined["prev"] > 0]
    combined["pct"] = combined.apply(
        lambda r: pct_change(r["prev"], r["curr"]), axis=1)
    gainers = combined.nlargest(5, "pct")
    losers  = combined.nsmallest(5, "pct")
    subset  = pd.concat([gainers, losers]).drop_duplicates()
    subset  = subset.sort_values("pct")

    colors = [C_UP if v >= 0 else C_DOWN for v in subset["pct"]]
    texts  = [f"{'▲' if v >= 0 else '▼'}{abs(v):.0f}%" for v in subset["pct"]]

    fig = go.Figure(go.Bar(
        y=subset.index.tolist(),
        x=subset["pct"].tolist(),
        orientation="h",
        marker=dict(color=colors, line=dict(width=0)),
        text=texts, textposition="outside",
        textfont=dict(size=11, family=FONT),
        hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
    ))
    fig.add_vline(x=0, line_width=1, line_color=C_GRID)
    fig.update_layout(**base_layout(
        title=dict(
            text=f"<b>Biggest Movers — {prev_lbl} vs {curr_lbl}</b><br><sup>{subtitle}</sup>",
            x=0.5, xanchor="center"),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="% Change",
                   ticksuffix="%"),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=11)),
        height=max(420, len(subset) * 45 + 120),
        margin=dict(t=100, b=60, l=180, r=80),
    ))
    return fig


def extract_daily_total(path, filename):
    """Read a daily 'Борлуулалт бараагаар' XLS and return (date, total_revenue).
    Date is extracted from the filename (pattern yyyymmdd).
    Total revenue is taken from the 'Дүн:' grand-total row (col index 2).
    """
    m = re.search(r"(\d{8})", filename)
    if not m:
        return None
    try:
        dt = pd.to_datetime(m.group(1), format="%Y%m%d")
    except Exception:
        return None
    try:
        raw = pd.read_excel(path, header=None, dtype=object)
        mask = raw.iloc[:, 0].astype(str).str.strip() == "Дүн:"
        dun = raw[mask]
        if dun.empty:
            return None
        # Revenue is at col 9 in the Дүн: grand-total row
        rev = dun.iloc[0, 9] if dun.shape[1] > 9 else None
        if rev is None or str(rev).strip() in ("", "nan"):
            return None
        return (dt, float(rev))
    except Exception:
        return None


def daily_files_vs_report_chart(daily_totals, report_daily):
    """Grouped bar chart: each day's total from individual daily report files
    vs the same day's total from the merged weekly report.
    daily_totals : list of (datetime, float)
    report_daily : dict  {"MM-DD": float}
    """
    DAY_ABBR = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
    daily_totals_sorted = sorted(daily_totals, key=lambda x: x[0])

    labels, vals_files, vals_report = [], [], []
    for dt, rev in daily_totals_sorted:
        md = dt.strftime("%m-%d")
        label = f"{DAY_ABBR[dt.weekday()]} {dt.month}/{dt.day}"
        labels.append(label)
        vals_files.append(rev)
        vals_report.append(float(report_daily.get(md, 0)))

    total_files  = sum(vals_files)
    total_report = sum(vals_report)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Daily Report Files", x=labels, y=vals_files,
        marker=dict(color=C_CURR, line=dict(width=0)),
        text=[fmt(v) for v in vals_files], textposition="outside",
        textfont=dict(size=10, color="#0f172a"),
        hovertemplate="<b>%{x}</b><br>Daily file: %{y:,.0f} ₮<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Weekly Report", x=labels, y=vals_report,
        marker=dict(color=C_PREV, line=dict(width=0)),
        text=[fmt(v) for v in vals_report], textposition="outside",
        textfont=dict(size=10, color="#475569"),
        hovertemplate="<b>%{x}</b><br>Weekly report: %{y:,.0f} ₮<extra></extra>",
    ))
    fig.update_layout(**base_layout(
        barmode="group",
        title=dict(
            text=(f"<b>Daily Revenue — Report Files vs Weekly Report</b><br>"
                  f"<sup>Files total: {fmt(total_files)} ₮  |  "
                  f"Weekly report total: {fmt(total_report)} ₮</sup>"),
            x=0.5, xanchor="center",
        ),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=12)),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="Revenue (₮)"),
        height=500,
        margin=dict(t=110, b=60, l=70, r=40),
        legend=dict(orientation="h", x=0.5, y=1.06, xanchor="center"),
    ))
    return fig


# ── plan category keyword map ────────────────────────────────────────────────────
PLAN_CAT_KW = {
    "bottled beer":   ["chimay","leffe","westmalle","bavik","lindemans","petrus","wittekerke",
                       "corona","heineken","stella","budweiser","0.33b","0.5b","0.33l","0.5l",
                       "bottle"],
    "draft beer":     ["draft","tap"],
    "guest beer":     ["guest beer","import beer"],
    "main course":    ["steak","salmon","pork belly","chicken leg","goulash","spareribs",
                       "fillet","t-bone","ox bone","bbq","carbonara","grill platter"],
    "share food":     ["share","platter"],
    "pub food":       ["burger","sandwich","pub food","wings"],
    "pizza":          ["pizza","margherita","meat lover","spicy salami"],
    "pasta":          ["pasta","penne"],
    "appetizer":      ["cobb","avocado","salad","appetizer","starter"],
    "side dish":      ["fries","french fries","side dish","sauce","bread","snack"],
    "soup":           ["soup","bone broth"],
    "dessert":        ["dessert","cake","brownie","ice cream","waffle"],
    "whiskey":        ["whiskey","whisky","jameson","scotch","bourbon","glenmorange"],
    "single malt":    ["single malt"],
    "vodka":          ["vodka","beluga","finlandia","grey goose","absolut","soyombo"],
    "gin":            ["gin","bombay","hendricks","gordons","gordon","tanqueray"],
    "cognac":         ["cognac","hennessy","brandy"],
    "tequila":        ["tequila","patron","jose cuervo","sierra"],
    "liquers":        ["liqueur","liquer","cointreau","baileys","schnapps","amaretto"],
    "water":          ["water","khujirt","millenia"],
    "soft drink":     ["cola","pepsi","sprite","fanta","schweppes","tonic","ginger beer",
                       "club soda","soda"],
    "tea":            ["tea","althaus"],
    "coffee":         ["coffee","espresso","latte","cappuccino","americano"],
    "juice":          ["juice","langers"],
    "milkshake":      ["milkshake","milk shake"],
    "hot drinks":     ["hot chocolate","hot drink"],
    "lunch set":      ["lunch","set menu","lunch set"],
    "white wine":     ["chardonnay","sauvignon","riesling","pinot grigio","cormons","moscato",
                       "trebbino","varaschin","poesie","banfi","white wine","terre forti"],
    "red wine":       ["merlot","shiraz","cabernet","pinot noir","bardolino","red wine"],
    "rose wine":      ["rose","rosé"],
    "champagne":      ["champagne","prosecco","moet","sparkling","cava"],
    "cocktails":      ["cocktail","mojito","margarita","martini","negroni","sour","highball",
                       "lemonade","punch","milkshake"],
    "set":            ["set ","set,"],
    "other/room fee": ["service charge","room fee","cover charge"],
}


def match_plan_category(item_name: str) -> str:
    """Match an item name to a sales plan category key."""
    n = item_name.lower()
    # Single malt before whiskey to avoid overlap
    for cat in ["single malt","draft beer","guest beer"]:
        for kw in PLAN_CAT_KW[cat]:
            if kw in n:
                return cat
    for cat, kws in PLAN_CAT_KW.items():
        for kw in kws:
            if kw in n:
                return cat
    return "other/room fee"


def fill_sales_plan_excel(report_df, plan_path):
    """
    Fill the Гүйцэтгэл columns in the sales plan with actual qty from report_df.
    Returns modified Excel as bytes.
    report_df must have columns: date (datetime), item_col (str), qty_col (numeric).
    """
    import copy

    # Detect item and qty columns
    date_col, item_col, qty_col, _, _, _ = detect_cols(report_df)
    if not item_col or not qty_col or not date_col:
        return None

    # Prepare report: parse dates, compute day-of-week, assign plan categories
    d = report_df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    d = d.dropna(subset=[date_col])
    d[qty_col]  = pd.to_numeric(d[qty_col],  errors="coerce").fillna(0)
    d["_date"]  = d[date_col].dt.normalize()
    d["_dow"]   = d[date_col].dt.day_name()
    d["_pcat"]  = d[item_col].fillna("").apply(match_plan_category)

    day_offsets = {
        "Monday": 0, "Tuesday": 9, "Wednesday": 18,
        "Thursday": 27, "Friday": 36, "Saturday": 45, "Sunday": 54,
    }
    # Mon/Tue/Sat/Sun: actual_col = offset+3, pct_col = offset+4
    # Wed/Thu/Fri:     actual_col = offset+2, pct_col = offset+3
    high_days = {"Wednesday","Thursday","Friday"}

    wb = openpyxl.load_workbook(plan_path)
    ws = wb.active

    # Read plan categories from Excel (row 4 onward = Excel row index 4..37 = 0-indexed 3..36)
    raw = pd.read_excel(plan_path, header=None, dtype=str)

    for day, off in day_offsets.items():
        day_data = d[d["_dow"] == day]
        actual_by_cat = day_data.groupby("_pcat")[qty_col].sum()

        is_high = day in high_days
        actual_col_off = 2 if is_high else 3   # relative to day offset
        pct_col_off    = 3 if is_high else 4

        for ri in range(3, 37):  # 0-indexed rows 3-36 = Excel rows 4-37
            plan_cat = str(raw.iloc[ri, off]).strip().lower()
            if not plan_cat or plan_cat == "nan":
                continue
            plan_qty_raw = raw.iloc[ri, off + 1]
            try:
                plan_qty = float(str(plan_qty_raw).replace(",", ""))
            except (ValueError, TypeError):
                plan_qty = 0.0

            actual_qty = float(actual_by_cat.get(plan_cat, 0))
            pct_val    = (actual_qty / plan_qty) if plan_qty else 0.0

            excel_row = ri + 1   # 0-indexed row ri → Excel row ri+1 (1-based)
            excel_col_actual = off + actual_col_off + 1   # +1 for 1-based Excel col
            excel_col_pct    = off + pct_col_off    + 1

            ws.cell(row=excel_row, column=excel_col_actual).value = round(actual_qty, 2)
            ws.cell(row=excel_row, column=excel_col_pct).value    = round(pct_val, 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_sales_plan(path):
    """
    Parse the weekly sales plan Excel file.
    Detects files containing 'борлуулалттай өдөр' in first 3 rows.
    Returns {"revenue": {day: target}, "qty": {day: total_qty_target}}
    or None if not a plan file.
    """
    try:
        raw = pd.read_excel(path, header=None, dtype=str)
    except Exception:
        return None

    flat = " ".join(str(v) for v in raw.iloc[:3].fillna("").values.flatten())
    if "борлуулалттай өдөр" not in flat:
        return None

    day_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    row0 = raw.iloc[0].tolist()
    day_offsets = {}
    for ci, v in enumerate(row0):
        if pd.notna(v):
            for d in day_order:
                if d.lower() in str(v).strip().lower():
                    day_offsets[d] = ci
                    break

    if not day_offsets:
        return None

    rev_plan = {}
    qty_plan = {}
    row1 = raw.iloc[1].tolist()

    for day, offset in day_offsets.items():
        # Revenue target from row 1
        cell = str(row1[offset]) if offset < len(row1) else ""
        nums = re.findall(r"[\d,]+\.?\d*", cell.replace(" ", ""))
        if nums:
            rev_plan[day] = float(nums[0].replace(",", ""))

        # Qty targets: col offset+1 for each category row (rows 3 onwards)
        total_qty = 0.0
        for ri in range(3, min(50, len(raw))):
            cat = str(raw.iloc[ri, offset]).strip() if pd.notna(raw.iloc[ri, offset]) else ""
            if not cat or cat in ("nan", ""):
                continue
            qty_cell = raw.iloc[ri, offset + 1] if offset + 1 < raw.shape[1] else None
            if qty_cell is not None and str(qty_cell).strip() not in ("", "nan"):
                try:
                    total_qty += float(str(qty_cell).replace(",", ""))
                except ValueError:
                    pass
        if total_qty > 0:
            qty_plan[day] = total_qty

    if not rev_plan:
        return None
    return {"revenue": rev_plan, "qty": qty_plan}


def _plan_daily_chart(labels, actuals, targets, title, subtitle, y_title, week_label):
    """Shared grouped bar chart: actual vs plan per day, green/red coloring."""
    bar_colors = [C_UP if a >= t else C_DOWN for a, t in zip(actuals, targets)]
    annotations = []
    for i, (a, t) in enumerate(zip(actuals, targets)):
        p = (a / t * 100) if t else 0
        arrow = "▲" if a >= t else "▼"
        col = C_UP if a >= t else C_DOWN
        annotations.append(dict(
            x=labels[i], y=max(a, t),
            xref="x", yref="y",
            text=f"<b>{arrow}{p:.0f}%</b>",
            showarrow=False, yshift=36,
            font=dict(size=12, family=FONT, color=col),
            align="center",
        ))

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Plan Target", x=labels, y=targets,
        marker=dict(color=C_PREV, line=dict(width=0)),
        text=[fmt(v) for v in targets], textposition="outside",
        textfont=dict(size=10, color="#475569"),
        hovertemplate="<b>%{x}</b><br>Target: %{y:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Actual", x=labels, y=actuals,
        marker=dict(color=bar_colors, line=dict(width=0)),
        text=[fmt(v) for v in actuals], textposition="outside",
        textfont=dict(size=10, color="#0f172a"),
        hovertemplate="<b>%{x}</b><br>Actual: %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(**base_layout(
        barmode="group",
        title=dict(
            text=f"<b>{title} — {week_label}</b><br><sup>{subtitle}</sup>",
            x=0.5, xanchor="center",
        ),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=12)),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title=y_title),
        annotations=annotations,
        height=500,
        margin=dict(t=110, b=60, l=70, r=40),
        legend=dict(orientation="h", x=0.5, y=1.06, xanchor="center"),
    ))
    return fig


def daily_vs_plan_chart(df, date_col, rev_col, qty_col, plan_data, week_label):
    """Two charts: daily revenue vs plan + daily qty vs plan."""
    d = df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    d = d.dropna(subset=[date_col])
    d["_date"] = d[date_col].dt.normalize()
    d["_dow"]  = d[date_col].dt.day_name()

    agg = {rev_col: "sum"}
    if qty_col and qty_col in d.columns:
        agg[qty_col] = "sum"
    daily = d.groupby(["_date","_dow"]).agg(agg).reset_index().sort_values("_date")

    labels = [f"{r['_dow'][:3]}  {str(r['_date'].date())[5:]}" for _, r in daily.iterrows()]

    rev_targets = plan_data.get("revenue", {})
    qty_targets = plan_data.get("qty", {})

    actuals_rev = [float(r[rev_col]) for _, r in daily.iterrows()]
    targets_rev = [float(rev_targets.get(r["_dow"], 0)) for _, r in daily.iterrows()]

    fig_rev = _plan_daily_chart(
        labels, actuals_rev, targets_rev,
        "Daily Revenue vs Plan",
        "Green = reached target  |  Red = below target",
        "Revenue (₮)", week_label,
    )

    fig_qty = None
    if qty_col and qty_col in daily.columns and qty_targets:
        actuals_qty = [float(r[qty_col]) for _, r in daily.iterrows()]
        targets_qty = [float(qty_targets.get(r["_dow"], 0)) for _, r in daily.iterrows()]
        fig_qty = _plan_daily_chart(
            labels, actuals_qty, targets_qty,
            "Daily Items Sold vs Break-Even Target",
            "Break-even qty = minimum items needed per day",
            "Items Sold", week_label,
        )

    return fig_rev, fig_qty


def weekly_vs_plan_chart(df, date_col, rev_col, qty_col, plan_data, week_label):
    """Weekly summary: actual vs plan for both revenue and qty."""
    d = df.copy()
    d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    d = d.dropna(subset=[date_col])
    d["_dow"] = d[date_col].dt.day_name()

    rev_targets = plan_data.get("revenue", {})
    qty_targets = plan_data.get("qty", {})

    days_in_data = d["_dow"].dropna().unique()
    actual_rev   = float(d[rev_col].sum())
    plan_rev     = sum(rev_targets.get(day, 0) for day in days_in_data)
    if plan_rev == 0:
        plan_rev = sum(rev_targets.values())

    actual_qty = float(d[qty_col].sum()) if qty_col and qty_col in d.columns else None
    plan_qty   = sum(qty_targets.get(day, 0) for day in days_in_data) if qty_targets else None
    if plan_qty == 0:
        plan_qty = sum(qty_targets.values()) if qty_targets else None

    def pct_label(actual, plan):
        if not plan:
            return ""
        p = actual / plan * 100
        arrow = "▲" if actual >= plan else "▼"
        col = C_UP if actual >= plan else C_DOWN
        return arrow, p, col

    # Revenue summary chart
    rev_arrow, rev_pct, rev_col_c = pct_label(actual_rev, plan_rev)
    fig_rev = go.Figure()
    fig_rev.add_trace(go.Bar(
        name="Weekly Plan", x=["Plan"], y=[plan_rev],
        marker=dict(color=C_PREV, line=dict(width=0)),
        text=[fmt(plan_rev)], textposition="outside",
        textfont=dict(size=14, color="#475569"),
        hovertemplate="Plan: %{y:,.0f}<extra></extra>",
    ))
    fig_rev.add_trace(go.Bar(
        name="Actual", x=["Actual"], y=[actual_rev],
        marker=dict(color=rev_col_c, line=dict(width=0)),
        text=[fmt(actual_rev)], textposition="outside",
        textfont=dict(size=14, color="#0f172a"),
        hovertemplate="Actual: %{y:,.0f}<extra></extra>",
    ))
    fig_rev.update_layout(**base_layout(
        barmode="group",
        title=dict(
            text=f"<b>Weekly Revenue vs Plan — {week_label}</b><br>"
                 f"<sup>{fmt(actual_rev)} ₮  of  {fmt(plan_rev)} ₮ target  "
                 f"({rev_pct:.1f}% achieved)</sup>",
            x=0.5, xanchor="center",
        ),
        xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=14)),
        yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="Revenue (₮)"),
        annotations=[dict(
            x=0.5, y=max(actual_rev, plan_rev),
            xref="paper", yref="y",
            text=f"<b>{rev_arrow} {rev_pct:.1f}% of plan</b>",
            showarrow=False, yshift=22,
            font=dict(size=16, family=FONT, color=rev_col_c),
            align="center",
        )],
        height=480,
        margin=dict(t=120, b=60, l=70, r=40),
        legend=dict(orientation="h", x=0.5, y=1.07, xanchor="center"),
    ))

    # Qty summary chart
    fig_qty = None
    if actual_qty is not None and plan_qty:
        qty_arrow, qty_pct, qty_col_c = pct_label(actual_qty, plan_qty)
        fig_qty = go.Figure()
        fig_qty.add_trace(go.Bar(
            name="Weekly Plan", x=["Plan"], y=[plan_qty],
            marker=dict(color=C_PREV, line=dict(width=0)),
            text=[f"{plan_qty:.0f} items"], textposition="outside",
            textfont=dict(size=14, color="#475569"),
            hovertemplate="Plan: %{y:,.0f} items<extra></extra>",
        ))
        fig_qty.add_trace(go.Bar(
            name="Actual", x=["Actual"], y=[actual_qty],
            marker=dict(color=qty_col_c, line=dict(width=0)),
            text=[f"{actual_qty:.0f} items"], textposition="outside",
            textfont=dict(size=14, color="#0f172a"),
            hovertemplate="Actual: %{y:,.0f} items<extra></extra>",
        ))
        fig_qty.update_layout(**base_layout(
            barmode="group",
            title=dict(
                text=f"<b>Weekly Items Sold vs Break-Even Target — {week_label}</b><br>"
                     f"<sup>{actual_qty:.0f} items sold  of  {plan_qty:.0f} items needed  "
                     f"({qty_pct:.1f}% achieved)</sup>",
                x=0.5, xanchor="center",
            ),
            xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=14)),
            yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="Items Sold"),
            annotations=[dict(
                x=0.5, y=max(actual_qty, plan_qty),
                xref="paper", yref="y",
                text=f"<b>{qty_arrow} {qty_pct:.1f}% of break-even</b>",
                showarrow=False, yshift=22,
                font=dict(size=16, family=FONT, color=qty_col_c),
                align="center",
            )],
            height=480,
            margin=dict(t=120, b=60, l=70, r=40),
            legend=dict(orientation="h", x=0.5, y=1.07, xanchor="center"),
        ))

    return fig_rev, fig_qty


# ── main pipeline ───────────────────────────────────────────────────────────────

def generate_all_charts(df, source_label="", plan_data=None, daily_totals=None, report_daily=None):
    charts = []

    # ── Detect if parsed report format (has clean column names) ──────────────
    is_parsed = "item" in df.columns and "category" in df.columns and "year" in df.columns

    if is_parsed:
        item_col = "item"
        cat_col  = "category"
        qty_col  = "qty"  if "qty"     in df.columns else None
        rev_col  = "revenue" if "revenue" in df.columns else None
        emp_col  = None
        date_col = None

        # YoY split from year column
        years = sorted(df["year"].dropna().unique().astype(int))
        has_yoy = len(years) >= 2
        if has_yoy:
            py, cy = years[-2], years[-1]
            prev_df = df[df["year"] == py].copy()
            curr_df = df[df["year"] == cy].copy()
            prev_lbl, curr_lbl = str(py), str(cy)
        else:
            cy = years[0] if years else 0
            prev_df = None
            curr_df = df.copy()
            prev_lbl, curr_lbl = "", str(cy)

        subtitle = f"{prev_lbl} vs {curr_lbl}" if has_yoy else curr_lbl

        # Categories come directly from the data — no keyword assignment needed
        curr_df["_sub_cat"], _ = zip(*curr_df[item_col].fillna("").map(assign_category))
        if has_yoy and prev_df is not None:
            prev_df["_sub_cat"], _ = zip(*prev_df[item_col].fillna("").map(assign_category))

    else:
        # Regular Excel format — use detect_cols
        date_col, item_col, qty_col, rev_col, emp_col, cat_col = detect_cols(df)
        if date_col:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
            # Drop separator rows and footer totals (invalid/out-of-range dates)
            valid_years = df[date_col].dt.year.between(1990, 2100)
            if valid_years.any():
                df = df[valid_years].reset_index(drop=True)

        if item_col and cat_col is None:
            df["_cat"], df["_sub"] = zip(*df[item_col].fillna("").map(assign_category))
            cat_col = "_cat"
        elif cat_col:
            df["_cat"] = df[cat_col]
        else:
            df["_cat"] = "All"
            cat_col = "_cat"
        df["_sub_cat"] = df["_cat"]

        prev_df, curr_df, prev_lbl, curr_lbl = None, df, "", "Current"
        has_yoy = False

        if date_col and df[date_col].notna().any():
            years = sorted(df[date_col].dt.year.dropna().unique().astype(int))
            if len(years) >= 2:
                py, cy = years[-2], years[-1]
                prev_df = df[df[date_col].dt.year == py].copy()
                curr_df = df[df[date_col].dt.year == cy].copy()
                prev_lbl, curr_lbl = str(py), str(cy)
                has_yoy = True
            else:
                curr_lbl = str(years[0])

        subtitle = f"{prev_lbl} vs {curr_lbl}" if has_yoy else curr_lbl

    # ── helpers ───────────────────────────────────────────────────────────────
    def grp(d, grp_col, val_col):
        if d is None or val_col is None or val_col not in d.columns:
            return pd.Series(dtype=float)
        return d.groupby(grp_col)[val_col].sum()

    def make_yoy(prev_s, curr_s, title, metric_label, orientation="v"):
        all_cats = list(curr_s.index.union(prev_s.index if prev_s is not None and not prev_s.empty else []))
        cv = [float(curr_s.get(c, 0)) for c in all_cats]
        pv = [float(prev_s.get(c, 0)) if prev_s is not None and not prev_s.empty else 0 for c in all_cats]
        fn = yoy_vertical if orientation == "v" else yoy_horizontal
        return fn(pv, cv, all_cats, title, subtitle, metric_label, prev_lbl, curr_lbl, source_label)

    # ── 1. Revenue overview by category ──────────────────────────────────────
    if rev_col and rev_col in df.columns:
        curr_cat_rev = grp(curr_df, cat_col, rev_col)
        prev_cat_rev = grp(prev_df, cat_col, rev_col) if has_yoy else pd.Series(dtype=float)
        curr_total   = curr_cat_rev.sum()
        prev_total   = prev_cat_rev.sum() if not prev_cat_rev.empty else 0
        overall_pct  = pct_change(prev_total, curr_total)
        sign = "▲" if overall_pct >= 0 else "▼"
        lbl  = source_label.upper() if source_label else "TOTAL"
        title_rev = (f"SALES REPORT — {lbl}  {sign}{abs(overall_pct):.1f}%"
                     f"  |  {fmt(prev_total)} -> {fmt(curr_total)} ₮")
        charts.append({"id": "rev_overview", "title": "Revenue by Category",
                        "fig": to_json(make_yoy(prev_cat_rev, curr_cat_rev, title_rev, "Revenue (₮)"))})

    # ── 2. Quantity overview by category ─────────────────────────────────────
    if qty_col and qty_col in df.columns:
        curr_cat_qty = grp(curr_df, cat_col, qty_col)
        prev_cat_qty = grp(prev_df, cat_col, qty_col) if has_yoy else pd.Series(dtype=float)
        if curr_cat_qty.sum() > 0:
            charts.append({"id": "qty_overview", "title": "Quantity by Category",
                            "fig": to_json(make_yoy(prev_cat_qty, curr_cat_qty,
                                                     f"QUANTITY OVERVIEW — {source_label or 'TOTAL'}",
                                                     "Quantity (pcs)"))})

    # ── 3. Staff revenue ──────────────────────────────────────────────────────
    if emp_col and rev_col and emp_col in df.columns:
        fig = staff_revenue_chart(curr_df, emp_col, rev_col, curr_lbl, subtitle, source_label)
        charts.append({"id": "staff_rev", "title": "Staff Revenue", "fig": to_json(fig)})

    # ── 4. Top 20 Items by Revenue ────────────────────────────────────────────
    if rev_col and item_col and item_col in curr_df.columns:
        curr_top_rev = grp(curr_df, item_col, rev_col).nlargest(20)
        prev_top_rev = (grp(prev_df, item_col, rev_col).reindex(curr_top_rev.index, fill_value=0)
                        if has_yoy and prev_df is not None else pd.Series(dtype=float))
        if curr_top_rev.sum() > 0:
            charts.append({"id": "top_items_rev", "title": "Top 20 Items by Revenue",
                "fig": to_json(make_yoy(prev_top_rev, curr_top_rev,
                    f"TOP 20 ITEMS — REVENUE  |  {source_label or subtitle}",
                    "Revenue (₮)", "h"))})

    # ── 5. Top 20 Items by Quantity ───────────────────────────────────────────
    if qty_col and item_col and item_col in curr_df.columns:
        curr_top_qty = grp(curr_df, item_col, qty_col).nlargest(20)
        prev_top_qty = (grp(prev_df, item_col, qty_col).reindex(curr_top_qty.index, fill_value=0)
                        if has_yoy and prev_df is not None else pd.Series(dtype=float))
        if curr_top_qty.sum() > 0:
            charts.append({"id": "top_items_qty", "title": "Top 20 Items by Quantity",
                "fig": to_json(make_yoy(prev_top_qty, curr_top_qty,
                    f"TOP 20 ITEMS — QUANTITY  |  {source_label or subtitle}",
                    "Quantity (pcs)", "h"))})

    # ── 6. Daily Revenue Trend (date-based only) ──────────────────────────────
    if date_col and date_col in df.columns and rev_col:
        try:
            charts.append({"id": "daily_trend", "title": "Daily Revenue Trend",
                "fig": to_json(daily_trend_chart(prev_df, curr_df, date_col, rev_col,
                                                  prev_lbl, curr_lbl, subtitle))})
        except Exception:
            pass

    # ── 7. Revenue by Day of Week ─────────────────────────────────────────────
    if date_col and date_col in curr_df.columns and rev_col:
        try:
            charts.append({"id": "dow", "title": "Revenue by Day of Week",
                "fig": to_json(dow_chart(curr_df, date_col, rev_col, curr_lbl, subtitle))})
        except Exception:
            pass

    # ── 8. Category Revenue Mix (donut) ──────────────────────────────────────
    if rev_col and cat_col and cat_col in curr_df.columns:
        cat_mix = grp(curr_df, cat_col, rev_col).nlargest(10)
        cat_mix = cat_mix[cat_mix > 0]
        if len(cat_mix) >= 2:
            charts.append({"id": "cat_donut", "title": "Revenue Mix by Category",
                "fig": to_json(donut_chart(
                    cat_mix.index.tolist(), cat_mix.values.tolist(),
                    f"REVENUE MIX — {curr_lbl}"))})

    # ── 9. Staff Category Mix (stacked 100% bar) ──────────────────────────────
    if (emp_col and cat_col and rev_col
            and emp_col in curr_df.columns and cat_col in curr_df.columns):
        try:
            pivot = (curr_df.groupby([emp_col, cat_col])[rev_col].sum()
                     .unstack(fill_value=0))
            pct = pivot.div(pivot.sum(axis=1), axis=0) * 100
            if pct.shape[0] >= 1:
                palette = ["#1a3c5e","#4a9fd4","#27ae60","#e67e22","#9b59b6",
                           "#e74c3c","#1abc9c","#f39c12","#2980b9","#d35400"]
                fig_mix = go.Figure()
                for idx, cat in enumerate(pct.columns):
                    fig_mix.add_trace(go.Bar(
                        name=str(cat),
                        x=pct.index.astype(str).tolist(),
                        y=pct[cat].tolist(),
                        marker=dict(color=palette[idx % len(palette)], line=dict(width=0)),
                        hovertemplate="<b>%{x}</b><br>" + str(cat) + ": %{y:.1f}%<extra></extra>",
                    ))
                fig_mix.update_layout(**base_layout(
                    barmode="stack",
                    title=dict(
                        text=f"<b>Staff Revenue Mix by Category</b><br><sup>{subtitle}</sup>",
                        x=0.5, xanchor="center"),
                    xaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, tickfont=dict(size=11)),
                    yaxis=dict(gridcolor=C_GRID, linecolor=C_GRID, title="% of Revenue",
                               ticksuffix="%", range=[0, 105]),
                    height=480,
                    legend=dict(orientation="v", x=1.01, y=0.99, xanchor="left"),
                ))
                charts.append({"id": "staff_mix", "title": "Staff Revenue Mix by Category",
                    "fig": to_json(fig_mix)})
        except Exception:
            pass

    # ── 10. YoY Biggest Movers ────────────────────────────────────────────────
    if has_yoy and rev_col and item_col and prev_df is not None:
        try:
            fig_mv = movers_chart(prev_df, curr_df, item_col, rev_col,
                                   prev_lbl, curr_lbl, subtitle)
            charts.append({"id": "movers", "title": "Biggest Movers (YoY % Change)",
                "fig": to_json(fig_mv)})
        except Exception:
            pass

    # ── 11-14. Sales Plan comparison (inserted at top) ───────────────────────
    if plan_data and date_col and date_col in curr_df.columns and rev_col:
        week_label = source_label or curr_lbl
        insert_pos = 0
        try:
            fig_rev_daily, fig_qty_daily = daily_vs_plan_chart(
                curr_df, date_col, rev_col, qty_col, plan_data, week_label)
            charts.insert(insert_pos, {"id": "daily_vs_plan",
                "title": "Daily Revenue vs Plan", "fig": to_json(fig_rev_daily)})
            insert_pos += 1
            if fig_qty_daily:
                charts.insert(insert_pos, {"id": "daily_qty_vs_plan",
                    "title": "Daily Items Sold vs Break-Even", "fig": to_json(fig_qty_daily)})
                insert_pos += 1
        except Exception:
            pass
        try:
            fig_rev_wk, fig_qty_wk = weekly_vs_plan_chart(
                curr_df, date_col, rev_col, qty_col, plan_data, week_label)
            charts.insert(insert_pos, {"id": "weekly_vs_plan",
                "title": "Weekly Revenue vs Plan", "fig": to_json(fig_rev_wk)})
            insert_pos += 1
            if fig_qty_wk:
                charts.insert(insert_pos, {"id": "weekly_qty_vs_plan",
                    "title": "Weekly Items vs Break-Even", "fig": to_json(fig_qty_wk)})
        except Exception:
            pass

    # ── 0. Daily comparison: individual files vs weekly report ────────────────
    if daily_totals and len(daily_totals) >= 2 and report_daily:
        try:
            fig_cmp = daily_files_vs_report_chart(daily_totals, report_daily)
            charts.insert(0, {
                "id": "daily_file_comparison",
                "title": "Daily Revenue: Report Files vs Weekly Report",
                "fig": to_json(fig_cmp),
            })
        except Exception:
            pass

    return charts


# ── routes ───────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400
    clear_uploads()
    saved = []
    for i, f in enumerate(files[:10]):
        if not f.filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
            continue
        buf = f.read()
        path = os.path.join(UPLOAD_FOLDER, f"upload_{i}.xlsx")
        with open(path, "wb") as fp:
            fp.write(buf)
        with open(path.replace(".xlsx", ".name"), "w") as fn:
            fn.write(f.filename)
        saved.append(f.filename)
    if not saved:
        return jsonify({"error": "No valid Excel files found"}), 400
    first = pd.ExcelFile(os.path.join(UPLOAD_FOLDER, "upload_0.xlsx"))
    return jsonify({"sheets": first.sheet_names, "files": saved})


@app.route("/api/preview", methods=["POST"])
def preview():
    data = request.get_json()
    sheet = data.get("sheet", 0)
    try:
        frames, _, _, _ = _load_files(sheet)
        df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        df = coerce_numerics(df)
        date_col, item_col, qty_col, rev_col, emp_col, _ = detect_cols(df)
        roles = {}
        for col in df.columns:
            s = df[col].dropna()
            if pd.api.types.is_datetime64_any_dtype(s):
                roles[col] = "date"
            elif col == date_col:
                roles[col] = "date"
            elif pd.api.types.is_numeric_dtype(s):
                roles[col] = "numeric"
            elif s.nunique() <= 50:
                roles[col] = "category"
            else:
                roles[col] = "text"
        detected = {k: v for k, v in {
            "date": date_col, "item": item_col,
            "qty": qty_col, "revenue": rev_col, "staff": emp_col
        }.items() if v}
        sample = df.head(5).fillna("").astype(str).to_dict("records")
        return jsonify({
            "columns": list(df.columns), "roles": roles,
            "shape": [int(df.shape[0]), int(df.shape[1])],
            "sample": sample, "detected": detected,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/charts", methods=["POST"])
def charts():
    data = request.get_json()
    sheet = data.get("sheet", 0)
    try:
        frames, plan_data, daily_totals, report_daily = _load_files(sheet)
        if not frames:
            return jsonify({"error": "No data loaded"}), 400

        # If multiple files, try to keep source labels per file
        if len(frames) == 1:
            df = frames[0]
            source = _source_label(0)
            df = coerce_numerics(df)
            result = generate_all_charts(df, source, plan_data, daily_totals, report_daily)
        else:
            # Merge all files — auto-detect years across them
            for i, f in enumerate(frames):
                f["_file_source"] = _source_label(i)
                coerce_numerics(f)
            df = pd.concat(frames, ignore_index=True)
            result = generate_all_charts(df, "", plan_data, daily_totals, report_daily)

        return jsonify({"charts": result})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


def parse_report_file(path):
    """
    Parse the Mongolian POS report format:
      - 'Бүлэг:' rows define category names
      - Product rows: col0=seq_num, col2=item_name, col4=year,
                      col7=qty, col9=revenue, col13=net_qty, col14=net_revenue
    Returns a clean DataFrame with columns: item, category, year, qty, revenue
    Or None if the file doesn't match this format.
    """
    try:
        raw = pd.read_excel(path, header=None, dtype=str)
    except Exception:
        return None

    # Detect if this is the report format
    flat = " ".join(raw.iloc[:15].fillna("").astype(str).values.flatten())
    if "Бүлэг:" not in flat and "Тайлант үе" not in flat:
        return None

    records = []
    current_cat = "Other"

    for _, row in raw.iterrows():
        c0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        c2 = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""

        # Category header row
        if c0 == "Бүлэг:":
            current_cat = c2 if c2 else current_cat
            continue

        # Product row: col 0 is a number
        if c0.isdigit() and c2:
            try:
                year_raw = str(row.iloc[4]).strip() if len(row) > 4 else ""
                year = int(float(year_raw)) if year_raw and year_raw not in ("nan", "") else None

                qty_raw = row.iloc[7] if len(row) > 7 else None
                qty = float(qty_raw) if qty_raw is not None and str(qty_raw) not in ("nan", "") else 0.0

                rev_raw = row.iloc[9] if len(row) > 9 else None
                rev = float(rev_raw) if rev_raw is not None and str(rev_raw) not in ("nan", "") else 0.0

                # Use net columns if sales are zero
                if qty == 0 and len(row) > 13:
                    net_qty_raw = row.iloc[13]
                    if net_qty_raw is not None and str(net_qty_raw) not in ("nan", ""):
                        qty = float(net_qty_raw)
                if rev == 0 and len(row) > 14:
                    net_rev_raw = row.iloc[14]
                    if net_rev_raw is not None and str(net_rev_raw) not in ("nan", ""):
                        rev = float(net_rev_raw)

                if qty > 0 or rev > 0:
                    records.append({
                        "item": c2,
                        "category": current_cat,
                        "year": year,
                        "qty": qty,
                        "revenue": rev,
                    })
            except (ValueError, IndexError):
                continue

    if not records:
        return None

    return pd.DataFrame(records)


def _load_files(sheet):
    frames = []
    plan_data = None
    daily_totals = []   # [(datetime, revenue)] — one entry per parsed daily report file
    report_daily = {}   # {"MM-DD": revenue}    — daily sums from a transaction-level report
    i = 0
    while True:
        p = os.path.join(UPLOAD_FOLDER, f"upload_{i}.xlsx")
        if not os.path.exists(p):
            break
        # Read original filename
        name_p = p.replace(".xlsx", ".name")
        filename = ""
        if os.path.exists(name_p):
            with open(name_p) as _fn:
                filename = _fn.read().strip()

        # Check if this is a sales plan file first
        plan = parse_sales_plan(p)
        if plan is not None:
            plan_data = plan
            i += 1
            continue
        # Try the structured report parser
        parsed = parse_report_file(p)
        if parsed is not None:
            frames.append(parsed)
            # Also extract the grand-total and date for cross-file comparison
            result = extract_daily_total(p, filename)
            if result:
                daily_totals.append(result)
        else:
            # Fallback: read as regular Excel
            try:
                xl = pd.ExcelFile(p)
                sname = xl.sheet_names[sheet] if isinstance(sheet, int) and sheet < len(xl.sheet_names) else xl.sheet_names[0]
                df = pd.read_excel(p, sheet_name=sname)
                df.columns = df.columns.astype(str).str.strip()
                df = df.dropna(how="all", axis=1).dropna(how="all", axis=0)
                # Remove separator/total rows (>80% of cells are empty)
                df = df[df.isnull().mean(axis=1) < 0.8].reset_index(drop=True)
                df = coerce_numerics(df)
                frames.append(df)
                # Compute daily revenue totals for cross-file comparison chart
                try:
                    dc, _, _, rc, _, _ = detect_cols(df)
                    if dc and rc and dc in df.columns and rc in df.columns:
                        tmp = df.copy()
                        tmp[dc] = pd.to_datetime(tmp[dc], errors="coerce")
                        tmp = tmp.dropna(subset=[dc])
                        tmp["_md"] = tmp[dc].dt.strftime("%m-%d")
                        for md, rev in tmp.groupby("_md")[rc].sum().items():
                            if md and str(md) != "NaT":
                                report_daily[md] = report_daily.get(md, 0) + float(rev)
                except Exception:
                    pass
            except Exception:
                pass
        i += 1
    return frames, plan_data, daily_totals, report_daily


def _source_label(i):
    p = os.path.join(UPLOAD_FOLDER, f"upload_{i}.xlsx")
    # read original filename from a sidecar if stored, else return empty
    name_p = p.replace(".xlsx", ".name")
    if os.path.exists(name_p):
        with open(name_p) as f:
            n = f.read().strip()
        # extract useful part e.g. "karaoke_2026.xlsx" → "Karaoke"
        base = os.path.splitext(n)[0]
        for kw in ["karaoke","restaurant","bar","cafe","kitchen"]:
            if kw in base.lower():
                return kw.capitalize()
    return ""


@app.route("/api/fill-plan", methods=["POST"])
def fill_plan():
    """Fill Гүйцэтгэл columns in the uploaded sales plan with actual qty from report."""
    try:
        # Find the plan file and report file among uploads
        plan_path   = None
        report_frames = []
        i = 0
        while True:
            p = os.path.join(UPLOAD_FOLDER, f"upload_{i}.xlsx")
            if not os.path.exists(p):
                break
            plan = parse_sales_plan(p)
            if plan is not None:
                plan_path = p
            else:
                parsed = parse_report_file(p)
                if parsed is None:
                    try:
                        xl = pd.ExcelFile(p)
                        df = pd.read_excel(p, sheet_name=xl.sheet_names[0])
                        df.columns = df.columns.astype(str).str.strip()
                        df = df[df.isnull().mean(axis=1) < 0.8].reset_index(drop=True)
                        df = coerce_numerics(df)
                        report_frames.append(df)
                    except Exception:
                        pass
                else:
                    report_frames.append(parsed)
            i += 1

        if not plan_path:
            return jsonify({"error": "No sales plan file found in uploads"}), 400
        if not report_frames:
            return jsonify({"error": "No report data found in uploads"}), 400

        report_df = pd.concat(report_frames, ignore_index=True) if len(report_frames) > 1 else report_frames[0]
        report_df = coerce_numerics(report_df)

        buf = fill_sales_plan_excel(report_df, plan_path)
        if buf is None:
            return jsonify({"error": "Could not fill plan — could not detect date/item/qty columns in report"}), 400

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="sales_plan_filled.xlsx",
        )
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


if __name__ == "__main__":
    print("=" * 60)
    print("  Brussels Report Maker")
    print("  Open http://localhost:8080 in your browser")
    print("=" * 60)
    app.run(debug=False, port=8080, host="0.0.0.0")
